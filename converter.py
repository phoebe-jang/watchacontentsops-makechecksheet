"""편성표 CSV → 검수시트 미리보기 변환 로직."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

DAYS = ["월", "화", "수", "목", "금", "토", "일"]
WEEKDAYS = ["월", "화", "수", "목", "금"]  # 검수시트가 만들어지는 요일 (토/일은 금요일에 합침)
DAY_LABEL = {d: d + "요일" for d in DAYS}
DAY_PREV = {"월": "일", "화": "월", "수": "화", "목": "수", "금": "목", "토": "금", "일": "토"}

# 2026년 한국 공휴일 (M/D). 휴일 자동 판정용 — 편성표 비고에 "(기존 X에피)"가 있더라도
# X 요일의 실제 날짜가 이 목록에 있을 때만 휴일로 인식한다.
KR_HOLIDAYS_2026 = {
    "1/1",       # 신정
    "2/16", "2/17", "2/18",  # 설날 연휴
    "3/1", "3/2",            # 삼일절(일) + 대체공휴일
    "5/1",       # 근로자의 날
    "5/5",       # 어린이날
    "5/24", "5/25",  # 부처님오신날(일) + 대체
    "6/6",       # 현충일
    "8/15", "8/17",  # 광복절(토) + 대체
    "9/24", "9/25", "9/26", "9/28",  # 추석 연휴 + 대체
    "10/3", "10/5",  # 개천절(토) + 대체
    "10/9",      # 한글날
    "12/25",     # 성탄절
}

# 검수시트 엑셀의 컬럼 헤더 (검수시트 예시 - 시트1.csv 1행 기준).
# 첫 컬럼은 빈 헤더(요일 라벨/종영 표시 들어가는 자리). 동일 이름 컬럼("담당자")이 중복 등장하지만
# 위치가 의미를 갖기 때문에 list 그대로 유지하고 인덱스로만 접근한다.
CHECKSHEET_HEADERS: list[str] = [
    "",                # 0  A열 (구분/요일/종영)
    "Tier",            # 1  B
    "원본검수",        # 2  C  ← 요일 헤더 행에서 신규 요약 텍스트가 들어가는 자리
    "담당자",          # 3  D
    "왓챠검수",        # 4  E
    "담당자",          # 5  F
    "series id",       # 6
    "series code",     # 7
    "series title",    # 8
    "season id",       # 9
    "season code",     # 10
    "season title",    # 11
    "mapping_type",    # 12
    "episode id",      # 13
    "episode code",    # 14
    "episode number",  # 15
    "formal_number",   # 16
    "episode title",   # 17
    "cp",              # 18
    "service_start",   # 19
    "service_end",     # 20
    "지원",
    "구분",
    "원소재화질",
    "서비스화질",
    "화질코멘트",
    "화질패널티",
    "성인물/키즈",
    "연령등급",
    "방통위/채널",
    "방통위/방영일",
    "영등위/관람등급번호",
    "주제",
    "선정성",
    "폭력성",
    "대사",
    "공포",
    "약물",
    "모방위험",
    "소재종류",
    "서비스 화면비",
    "화면비 비고",
    "음향",
    "자막",
    "더빙",
    "파일경로",        # 45
    "start offset(ms)",
    "end offset(ms)",
    "SoundTrack ( | 로 구분)",
    "VolumeOffset",
    "SoundLanguage ( | 로 구분)",
    "SDAR (교정화면비)",
    "SrcBucket",
    "비기닝(시간)",
    "엔딩(시간)",
    "beginning_sec",
    "opening_beginning_sec",
    "opening_ending_sec",
    "ending_sec",
    "nuvus_id",
    "비고 ",
    "백드롭",
    "짧은줄거리",
    "시즌 가로th",
    "시즌 세로th",
    "시즌 로고",
    "시리즈 가로th",
    "시리즈 세로th",
    "시리즈 로고",
    "담당자",          # 69 (마지막)
]

# 검수시트 컬럼 인덱스 상수
COL_A = 0
COL_TIER = 1
COL_NEW_SUMMARY = 2  # C열 - 요일 헤더 행에서 신규 요약 텍스트 위치
COL_SEASON_ID = 9
COL_SEASON_CODE = 10
COL_SEASON_TITLE = 11
COL_MAPPING_TYPE = 12
COL_EPISODE_NUMBER = 15
COL_FORMAL_NUMBER = 16
COL_CP = 18
COL_FILE_PATH = 45

# 미리보기에 표시할 컬럼 인덱스 (실제로 데이터가 채워지는 자리만).
PREVIEW_COL_INDICES = [
    COL_A,
    COL_TIER,
    COL_NEW_SUMMARY,
    COL_SEASON_ID,
    COL_SEASON_CODE,
    COL_SEASON_TITLE,
    COL_MAPPING_TYPE,
    COL_EPISODE_NUMBER,
    COL_FORMAL_NUMBER,
    COL_CP,
    COL_FILE_PATH,
]

# 미리보기 그리드 고정용 컬럼 폭 (px).
PREVIEW_COL_WIDTHS = {
    COL_A: 96,  # 구분 — 짧게. 긴 텍스트(요일 헤더/연휴지연 헤더)는 옆 빈 셀 위로 흘러나옴
    COL_TIER: 56,
    COL_NEW_SUMMARY: 200,
    COL_SEASON_ID: 100,
    COL_SEASON_CODE: 110,
    COL_SEASON_TITLE: 240,
    COL_MAPPING_TYPE: 96,
    COL_EPISODE_NUMBER: 90,
    COL_FORMAL_NUMBER: 90,
    COL_CP: 140,
    COL_FILE_PATH: 320,
}


def _empty_row() -> list[str]:
    return [""] * len(CHECKSHEET_HEADERS)


def parse_csv(file: Any) -> pd.DataFrame:
    df = pd.read_csv(file, dtype=str).fillna("")
    df.columns = [str(c).strip() for c in df.columns]
    return df


# 자막/더빙 꼬리 패턴: 하이픈/괄호/대괄호 + 공백 변형 모두 흡수
_SUB_SUFFIX_RE = re.compile(r"(?:\s*-\s*자막\s*|\s*\(\s*자막\s*\)\s*|\s*\[\s*자막\s*\]\s*)$")
_DUB_SUFFIX_RE = re.compile(r"(?:\s*-\s*더빙\s*|\s*\(\s*더빙\s*\)\s*|\s*\[\s*더빙\s*\]\s*)$")


def _strip_mapping_suffix(title: str) -> tuple[str, str]:
    """타이틀 끝의 자막/더빙 꼬리를 제거하고 mapping_type 반환.
    지원 패턴: ` - 더빙`, `-더빙`, ` (더빙)`, `[더빙]` 및 공백 변형 (자막도 동일)."""
    if _DUB_SUFFIX_RE.search(title):
        return _DUB_SUFFIX_RE.sub("", title).strip(), "dubbing"
    if _SUB_SUFFIX_RE.search(title):
        return _SUB_SUFFIX_RE.sub("", title).strip(), "basic"
    return title.strip(), ""


def _is_irregular(row: pd.Series) -> bool:
    yoil = str(row.get("요일", "") or "").strip()
    sigan = str(row.get("시간", "") or "").strip()
    return yoil == "비정기" or sigan == "비정기"


def _is_friday_ebs_special(row: pd.Series) -> bool:
    cp = str(row.get("CP bill", "") or "")
    yoil = str(row.get("요일", "") or "").strip()
    return cp.startswith("EBS") and yoil == ""


def _classify(row: pd.Series) -> str:
    bigo = str(row.get("편성비고", "") or "")
    bigo_a = str(row.get("비고", "") or "")
    e_n = str(row.get("이번주 e/n", "") or "").strip()
    sigan = str(row.get("시간", "") or "").strip()
    # 결방/홀드백: `이번주 e/n` 컬럼 값이 정확히 그 텍스트일 때만 (편성비고 부분 매칭 제거 — 다음 행 오인식 방지)
    if e_n in {"결방", "홀드백"}:
        return "cancelled"
    if "신규" in bigo:
        return "new"
    if "전날 셋팅" in bigo_a or "전날셋팅" in bigo_a:
        return "prev_day"
    if sigan and sigan != "18:00":
        return "reserved"
    return "normal"


def _parse_hour(sigan: str) -> int:
    m = re.match(r"^(\d{1,2}):", sigan.strip())
    return int(m.group(1)) if m else -1


def _display_day(row: pd.Series, prev_normal_day: str | None = None) -> str:
    yoil = str(row.get("요일", "") or "").strip()
    sigan = str(row.get("시간", "") or "").strip()
    bigo = str(row.get("비고", "") or "").strip()
    # 연휴지연편성: 편성표 csv 위치상 직전의 일반 요일에 따라가도록
    if "연휴지연" in bigo:
        result = prev_normal_day or yoil
    # 전날 셋팅: 강제로 전날
    elif "전날 셋팅" in bigo or "전날셋팅" in bigo:
        result = DAY_PREV.get(yoil, yoil)
    else:
        # 예약작 시간 < 12 → 전날
        cat = _classify(row)
        if cat == "reserved":
            h = _parse_hour(sigan)
            if 0 <= h < 12:
                result = DAY_PREV.get(yoil, yoil)
            else:
                result = yoil
        else:
            result = yoil
    # 토/일은 검수시트 없음 → 금요일로 합산 (사용자 룰)
    if result in ("토", "일"):
        result = "금"
    return result


def _compute_holiday_info(df: pd.DataFrame) -> tuple[set[str], dict[str, str]]:
    """편성표 비고에 '(기존 X에피)' 표기로 휴일 후보 추출 → 그 X 요일의 실제 날짜가
    `KR_HOLIDAYS_2026`에 있을 때만 휴일로 인정.
    Returns (휴일 요일 set, 요일 → 'M/D' 매핑)."""
    holiday_days: set[str] = set()
    day_to_date: dict[str, str] = {}
    pat = re.compile(r"\(\s*기존\s*([월화수목금토일])\s*에피\s*\)")

    candidate_days: set[str] = set()
    for _, row in df.iterrows():
        bigo = str(row.get("비고", "") or "")
        if "연휴지연" not in bigo:
            continue
        m = pat.search(bigo)
        if m:
            candidate_days.add(m.group(1))

    if not candidate_days:
        return holiday_days, day_to_date

    if "요일" in df.columns and "오픈일" in df.columns:
        for d in candidate_days:
            rows = df[df["요일"].astype(str).str.strip() == d]
            for _, row in rows.iterrows():
                opendate = str(row.get("오픈일", "") or "").strip()
                if not opendate:
                    continue
                nums = re.findall(r"\d+", opendate)
                if len(nums) >= 3:
                    md = f"{int(nums[1])}/{int(nums[2])}"
                elif len(nums) == 2:
                    md = f"{int(nums[0])}/{int(nums[1])}"
                else:
                    continue
                # 실제 한국 공휴일일 때만 휴일로 인정
                if md in KR_HOLIDAYS_2026:
                    holiday_days.add(d)
                    day_to_date[d] = md
                    print(
                        f"[휴일 판정] {d}요일 ({md}) → 공휴일 인정"
                    )
                else:
                    print(
                        f"[휴일 판정] {d}요일 ({md}) → 공휴일 아님, 일반 헤더 처리"
                    )
                break
    return holiday_days, day_to_date


def _compute_prev_normal_days(df: pd.DataFrame) -> list[str | None]:
    """편성표 csv 순서대로 각 행 직전의 '일반 요일'(연휴지연편성 X) 추적."""
    out: list[str | None] = []
    last: str | None = None
    for _, row in df.iterrows():
        bigo = str(row.get("비고", "") or "").strip()
        yoil = str(row.get("요일", "") or "").strip()
        if "연휴지연" not in bigo and yoil in DAYS:
            last = yoil
        out.append(last)
    return out


def check_jongyeong_alerts(df: pd.DataFrame) -> list[dict]:
    """`최종회차`보다 `이번주 e/n`이 더 큰 행 → 미고지 종영 가능성 알림.

    이미 편성비고에 '종영'이 적힌 행은 이미 처리됐다고 보고 알림 X.
    """
    alerts: list[dict] = []
    if "최종회차" not in df.columns:
        return alerts
    for _, row in df.iterrows():
        bigo = str(row.get("편성비고", "") or "")
        if "종영" in bigo:
            continue
        final_str = str(row.get("최종회차", "") or "")
        nums = re.findall(r"\d+", final_str)
        if not nums:
            continue
        # "13 (38-50화)" 같은 표기에서 마지막 회차 번호는 가장 큰 숫자(50)
        final_ep = max(int(n) for n in nums)
        e_n_str = str(row.get("이번주 e/n", "") or "")
        e_nums = re.findall(r"\d+", e_n_str)
        if not e_nums:
            continue
        max_e = max(int(n) for n in e_nums)
        if max_e > final_ep:
            title_raw = str(row.get("Title", "") or "").strip()
            title_clean, _mt = _strip_mapping_suffix(title_raw)
            alerts.append(
                {
                    "title": title_clean,
                    "final": final_ep,
                    "current_max": max_e,
                    "message": f"확인 필요 {title_clean} {final_ep}회차를 끝으로 종영인 것으로 확인됨",
                }
            )
    return alerts


_CP_SUFFIX_RE = re.compile(r"-[A-Z0-9]+$")  # 권리사 꼬리 코드 (예: -RS50, -RS65, -FLAT)


def _strip_cp_suffix(cp: str) -> str:
    """`CP bill`의 꼬리 코드 제거: 'WAG-RS50' → 'WAG', 'BELLMEDIA-FLAT' → 'BELLMEDIA'."""
    return _CP_SUFFIX_RE.sub("", str(cp or "").strip()).strip()


def _format_new_summary(new_items: list[dict]) -> str:
    """`신규: 타이틀[(자막/더빙)] e/n화 [(f/n N)] (권리사) / ...` 형식.
    - mapping_type이 basic이면 "(자막)", dubbing이면 "(더빙)"을 타이틀 바로 뒤(공백 없이)에 붙임.
    - f/n은 e/n과 다를 때만 `(f/n N)`으로 추가.
    - 권리사는 -RS 코드 제거된 형태."""
    if not new_items:
        return ""
    parts = []
    for r in new_items:
        title = r["Title"].strip()
        mt = r.get("mapping_type", "")
        type_label = "(더빙)" if mt == "dubbing" else "(자막)" if mt == "basic" else ""
        e_n = str(r.get("이번주 e/n", "") or "").strip()
        f_n = str(r.get("이번주 f/n", "") or "").strip()
        cp_clean = _strip_cp_suffix(r.get("CP bill", ""))

        # 타이틀 + (자막/더빙) — 공백 없음
        item = f"{title}{type_label}"
        # + 회차 — 공백 있음
        if e_n:
            item = f"{item} {e_n}화"
        # f/n 조건: 둘 다 있고 다를 때만
        show_fn = bool(e_n) and bool(f_n) and e_n != f_n
        if show_fn:
            item = f"{item} (f/n {f_n})"
        if cp_clean:
            item = f"{item} ({cp_clean})"
        parts.append(item)
        print(
            f"[신규 처리] raw={r['Title']!r} mt={mt!r} type_label={type_label!r} "
            f"en={e_n!r} fn={f_n!r} show_fn={show_fn} final={item!r}"
        )
    return "신규: " + " / ".join(parts)


def _split_episode_pairs(e_n: str, f_n: str) -> list[tuple[str, str]]:
    """`이번주 e/n`이 'X-Y' 범위면 X, X+1, ..., Y로 분할하고, f/n도 같은 개수로 짝지음."""
    e = str(e_n).strip()
    f = str(f_n).strip()
    m = re.match(r"^(\d+)-(\d+)$", e)
    if not m:
        return [(e, f)]
    a, b = int(m.group(1)), int(m.group(2))
    if b < a or (b - a) > 60:
        return [(e, f)]
    e_list = list(range(a, b + 1))
    fm = re.match(r"^(\d+)-(\d+)$", f)
    if fm:
        fa, fb = int(fm.group(1)), int(fm.group(2))
        f_list = list(range(fa, fb + 1))
        if len(f_list) != len(e_list):
            f_list = [""] * len(e_list)
    elif f == "":
        f_list = [""] * len(e_list)
    else:
        f_list = [f] + [""] * (len(e_list) - 1)
    return [(str(x), str(y)) for x, y in zip(e_list, f_list)]


def _calc_over_final(row: pd.Series, e_val: str, already_jongyeong: bool) -> bool:
    """이 분할 회차가 편성표 `최종회차` 숫자보다 큰지 (이미 '종영' 처리된 행은 제외)."""
    if already_jongyeong:
        return False
    final_str = str(row.get("최종회차", "") or "")
    final_nums = re.findall(r"\d+", final_str)
    if not final_nums:
        return False
    final_ep_val = max(int(n) for n in final_nums)
    e_nums = re.findall(r"\d+", str(e_val))
    if not e_nums:
        return False
    return max(int(n) for n in e_nums) > final_ep_val


def _row_values_for(row: pd.Series, e_val: str, f_val: str, jongyeong: bool) -> list[str]:
    vals = _empty_row()
    title_raw = str(row.get("Title", "") or "")
    title_clean, mt = _strip_mapping_suffix(title_raw)
    if jongyeong:
        vals[COL_A] = "종영"
    vals[COL_TIER] = str(row.get("Tier", "") or "")
    vals[COL_SEASON_ID] = str(row.get("id", "") or "")
    vals[COL_SEASON_CODE] = str(row.get("code", "") or "")
    vals[COL_SEASON_TITLE] = title_clean
    vals[COL_MAPPING_TYPE] = mt
    vals[COL_EPISODE_NUMBER] = e_val
    vals[COL_FORMAL_NUMBER] = f_val
    vals[COL_CP] = str(row.get("CP bill", "") or "")
    vals[COL_FILE_PATH] = str(row.get("원본경로", "") or "")
    return vals


def build_for_day(df: pd.DataFrame, target_day: str) -> dict:
    """선택된 요일 한 개에 대해 검수시트 미리보기 데이터 구조 반환."""
    df = df.copy()
    if "요일" not in df.columns:
        raise ValueError("CSV에 '요일' 컬럼이 없습니다. 헤더를 확인해주세요.")

    df = df[~df.apply(_is_irregular, axis=1)].reset_index(drop=True)

    ebs_mask = df.apply(_is_friday_ebs_special, axis=1)
    ebs_df = df[ebs_mask].copy().reset_index(drop=True)
    df = df[~ebs_mask].copy().reset_index(drop=True)

    prev_normal_days = _compute_prev_normal_days(df)
    holiday_days, day_to_date = _compute_holiday_info(df)
    is_target_holiday = target_day in holiday_days

    new_items: list[dict] = []
    # 편성표 csv 순서 그대로 보존 (예약작 정렬 X — 사용자 룰)
    mixed_rows: list[dict] = []  # reserved + normal을 csv 순서대로
    # 연휴지연편성: 비고 텍스트별로 그룹화 → 출력 시 한 헤더 + 그 아래 콘텐츠
    holiday_groups: dict[str, list[dict]] = {}
    holiday_group_order: list[str] = []
    prev_day_rows: list[dict] = []  # 비고 '전날 셋팅' 행 — 전날 맨 아래(결방 위)
    cancelled: list[dict] = []

    for idx, row in df.iterrows():
        cat = _classify(row)
        is_jongyeong = "종영" in str(row.get("편성비고", "") or "")
        bigo = str(row.get("비고", "") or "").strip()
        is_holiday = "연휴지연" in bigo
        prev = prev_normal_days[idx] if idx < len(prev_normal_days) else None
        display_d = _display_day(row, prev_normal_day=prev)

        # 월요일 행 + 비고 "전날 셋팅" → 전주 금요일에 이미 포함된 것이므로 이번 주 검수시트에서 제외
        if cat == "prev_day":
            yoil_orig = str(row.get("요일", "") or "").strip()
            if yoil_orig == "월":
                continue

        if cat == "new":
            if display_d != target_day:
                continue
            title_raw = str(row.get("Title", "") or "")
            title_clean, mt = _strip_mapping_suffix(title_raw)
            new_items.append(
                {
                    "Title": title_clean,
                    "이번주 e/n": str(row.get("이번주 e/n", "") or ""),
                    "이번주 f/n": str(row.get("이번주 f/n", "") or ""),
                    "CP bill": str(row.get("CP bill", "") or ""),
                    "mapping_type": mt,  # basic=자막 / dubbing=더빙 / ""=일반
                }
            )
            continue

        if display_d != target_day:
            continue

        e_n = str(row.get("이번주 e/n", "") or "")
        f_n = str(row.get("이번주 f/n", "") or "")
        pairs = _split_episode_pairs(e_n, f_n)

        if is_holiday:
            # 비고 텍스트별로 그룹 헤더 + 콘텐츠 두 톤 회색
            if bigo not in holiday_groups:
                holiday_groups[bigo] = []
                holiday_group_order.append(bigo)
            is_split = len(pairs) > 1
            group_id = f"holi-{idx}" if is_split else None
            for i, (e_val, f_val) in enumerate(pairs):
                jongyeong_here = is_jongyeong and i == len(pairs) - 1
                vals = _row_values_for(row, e_val, f_val, False)
                vals[COL_A] = ""  # 콘텐츠 행 A열은 비워둠 (헤더 행에만 비고 텍스트)
                holiday_groups[bigo].append(
                    {
                        "values": vals,
                        "category": "holiday_body",
                        "time": str(row.get("시간", "") or ""),
                        "jongyeong": False,
                        "group_id": group_id,
                        "split_idx": i,
                        "split_total": len(pairs),
                        "over_final": _calc_over_final(row, e_val, is_jongyeong),
                    }
                )
            continue

        is_split = len(pairs) > 1
        group_id = f"grp-{idx}" if is_split else None
        for i, (e_val, f_val) in enumerate(pairs):
            jongyeong_here = is_jongyeong and i == len(pairs) - 1
            vals = _row_values_for(row, e_val, f_val, jongyeong_here)
            # 전날 셋팅 행: A열에 원래 요일 표기 (예: "화 에피") — 예약작과 같은 핑크 음영
            if cat == "prev_day":
                yoil_orig = str(row.get("요일", "") or "").strip()
                if yoil_orig:
                    vals[COL_A] = f"{yoil_orig} 에피"
            entry = {
                "values": vals,
                "category": cat,
                "time": str(row.get("시간", "") or ""),
                "jongyeong": jongyeong_here,
                "group_id": group_id,
                "split_idx": i,
                "split_total": len(pairs),
                "over_final": _calc_over_final(row, e_val, is_jongyeong),
            }
            if cat == "cancelled":
                cancelled.append(entry)
            elif cat == "prev_day":
                prev_day_rows.append(entry)
            else:
                # reserved / normal 모두 편성표 csv 순서 그대로
                mixed_rows.append(entry)

    # 출력 순서: 편성표 csv 순서(예약·일반 섞임) → 전날 셋팅(결방 위)
    # (결방 → 연휴지연 그룹은 build_xlsx / 미리보기에서 cancelled 뒤에 별도로 이어 붙임)
    body: list[dict] = mixed_rows + prev_day_rows

    # 연휴지연 그룹 — 비고 텍스트별로 (헤더 1행 + 콘텐츠들)
    holiday_block: list[dict] = []
    for bigo_text in holiday_group_order:
        h_vals = _empty_row()
        h_vals[COL_A] = bigo_text
        holiday_block.append(
            {
                "values": h_vals,
                "category": "holiday_header",
                "time": "",
                "jongyeong": False,
                "over_final": False,
            }
        )
        holiday_block.extend(holiday_groups[bigo_text])

    # 디버깅 로그 (Streamlit 터미널에 출력)
    if holiday_block:
        print(
            f"[연휴지연 그룹 진입] target_day={target_day} "
            f"is_target_holiday={is_target_holiday} "
            f"groups={list(holiday_group_order)} "
            f"holiday_block_size={len(holiday_block)}"
        )
        for h in holiday_block[:5]:
            print(
                f"  - cat={h['category']} A={h['values'][COL_A]!r} "
                f"title={h['values'][COL_SEASON_TITLE]!r}"
            )

    # 휴일 요일이면 일반 콘텐츠만 음영 X (요일 헤더만 형광 초록).
    # 단, 결방(cancelled)은 어느 요일이든 빨간 음영 유지, 연휴지연(holiday_block)도 자체 회색 유지.
    if is_target_holiday:
        for entry in body:
            entry["no_fill"] = True
    # 결방 행 음영 적용 검증용 디버그 로그
    for entry in cancelled:
        title_d = entry["values"][COL_SEASON_TITLE]
        print(
            f"[결방 음영] target_day={target_day} category={entry['category']!r} "
            f"no_fill={entry.get('no_fill', False)} title={title_d!r}"
        )

    # 요일 헤더 행
    header_vals = _empty_row()
    if is_target_holiday:
        date_str = day_to_date.get(target_day, "")
        if date_str:
            header_vals[COL_A] = f"{DAY_LABEL[target_day]} ({date_str} 휴일)"
        else:
            header_vals[COL_A] = f"{DAY_LABEL[target_day]} (휴일)"
    else:
        header_vals[COL_A] = DAY_LABEL.get(target_day, target_day)
    header_vals[COL_NEW_SUMMARY] = _format_new_summary(new_items)
    header_row = {
        "values": header_vals,
        "category": "header_holiday" if is_target_holiday else "header",
        "new_count": len(new_items),
    }

    # EBS 섹션 (금요일 한정)
    ebs_section: list[dict] = []
    if target_day == "금":
        for _, row in ebs_df.iterrows():
            title_raw = str(row.get("Title", "") or "")
            title_clean, mt = _strip_mapping_suffix(title_raw)
            vals = _empty_row()
            vals[COL_A] = str(row.get("편성비고", "") or "")
            vals[COL_TIER] = str(row.get("Tier", "") or "")
            vals[COL_SEASON_ID] = str(row.get("id", "") or "")
            vals[COL_SEASON_CODE] = str(row.get("code", "") or "")
            vals[COL_SEASON_TITLE] = title_clean
            vals[COL_MAPPING_TYPE] = mt
            vals[COL_CP] = str(row.get("CP bill", "") or "")
            # 원본경로/episode number/formal_number 빈칸 (스펙)
            ebs_section.append({"values": vals, "category": "ebs"})

    return {
        "headers": CHECKSHEET_HEADERS,
        "header_row": header_row,
        "body": body,
        "cancelled": cancelled,
        "holiday_block": holiday_block,
        "ebs_section": ebs_section,
        "new_count": len(new_items),
    }


# ── XLSX 출력 ────────────────────────────────────────────────────────────

# 행 카테고리별 음영 (구글 시트 검수시트와 자연스럽게 어울리는 톤)
# 검수시트 템플릿 색 톤에 맞춘 fill (검수시트 예시 템플릿.xlsx 분석 결과)
_FILL_HEADER_DAY = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")      # 요일 헤더 회색
_FILL_RESERVED = PatternFill(start_color="FFEAD1DC", end_color="FFEAD1DC", fill_type="solid")        # 예약작 핑크 (시트2 가이드)
_FILL_CANCELLED = PatternFill(start_color="FFE6B8AF", end_color="FFE6B8AF", fill_type="solid")       # 결방/홀드백 빨간색 (시트2 가이드)
_FILL_EBS_BANNER = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")      # EBS 시작행 = 요일 헤더와 같은 회색
# (EBS 콘텐츠 행은 음영 X — _fill_for에서 None 반환)
_FILL_JONGYEONG = PatternFill(start_color="FFFFF292", end_color="FFFFF292", fill_type="solid")       # 종영 셀 노랑
_FILL_HOLIDAY_HEADER = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")  # 연휴지연 그룹 헤더 (요일과 같은 회색)
_FILL_HOLIDAY_BODY = PatternFill(start_color="FFB7B7B7", end_color="FFB7B7B7", fill_type="solid")    # 연휴지연 콘텐츠 (시트2 가이드 진한 회색)
_FILL_HEADER_HOLIDAY = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # 휴일 요일 헤더 (형광 초록)
_FILL_PREV_DAY = PatternFill(start_color="FFEAD1DC", end_color="FFEAD1DC", fill_type="solid")        # 전날 셋팅 = 예약작과 동일 핑크 (시트2 가이드)
_FILL_OVER_FINAL = PatternFill(start_color="FFCFE2F3", end_color="FFCFE2F3", fill_type="solid")      # 종영 의심 (옅은 파랑)
_FILL_ALERT = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")           # 알림(옅은 노랑)

# 폰트 — 템플릿: Arial 10pt
_FONT_BASE = Font(name="Arial", size=10)
_FONT_BOLD = Font(name="Arial", size=10, bold=True)
_FONT_BOLD_RED = Font(name="Arial", size=10, bold=True, color="FFFF0000")  # 연휴지연 헤더 강조용

# === 구글 시트 호환 모드 색상 (사용자 매핑) ===
# 예약작=연한자홍색3 / 결방=연한붉은딸기색3 / 요일&EBS=연한회색1 / 연휴지연 헤더=회색 /
# 연휴지연 콘텐츠=진한회색1 / 휴일=녹색
_GS_FILL_HEADER_DAY = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
_GS_FILL_HEADER_HOLIDAY = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
_GS_FILL_RESERVED = PatternFill(start_color="FFEAD1DC", end_color="FFEAD1DC", fill_type="solid")
_GS_FILL_CANCELLED = PatternFill(start_color="FFE6B8AF", end_color="FFE6B8AF", fill_type="solid")
_GS_FILL_PREV_DAY = PatternFill(start_color="FFEAD1DC", end_color="FFEAD1DC", fill_type="solid")
_GS_FILL_HOLIDAY_HEADER = PatternFill(start_color="FFB7B7B7", end_color="FFB7B7B7", fill_type="solid")
_GS_FILL_HOLIDAY_BODY = PatternFill(start_color="FF999999", end_color="FF999999", fill_type="solid")
_GS_FILL_EBS_BANNER = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
_GS_FILL_OVER_FINAL = PatternFill(start_color="FFCFE2F3", end_color="FFCFE2F3", fill_type="solid")


def _gs_fill_for(category: str):
    if category == "header":
        return _GS_FILL_HEADER_DAY
    if category == "header_holiday":
        return _GS_FILL_HEADER_HOLIDAY
    if category == "reserved":
        return _GS_FILL_RESERVED
    if category == "cancelled":
        return _GS_FILL_CANCELLED
    if category == "ebs":
        return None
    if category == "holiday_header":
        return _GS_FILL_HOLIDAY_HEADER
    if category == "holiday_body":
        return _GS_FILL_HOLIDAY_BODY
    if category == "prev_day":
        return _GS_FILL_PREV_DAY
    return None


def _fill_for(category: str) -> PatternFill | None:
    if category == "header":
        return _FILL_HEADER_DAY
    if category == "reserved":
        return _FILL_RESERVED
    if category == "cancelled":
        return _FILL_CANCELLED
    if category == "ebs":
        return None  # EBS 콘텐츠 행은 음영 없음 (배너만 회색)
    if category == "holiday_header":
        return _FILL_HOLIDAY_HEADER
    if category == "holiday_body":
        return _FILL_HOLIDAY_BODY
    if category == "header_holiday":
        return _FILL_HEADER_HOLIDAY
    if category == "prev_day":
        return _FILL_PREV_DAY
    return None


def build_ordered_entries(df: pd.DataFrame, days: list[str] | str) -> list[dict]:
    """미리보기/xlsx에 출력될 순서대로 평탄화된 엔트리 리스트.
    각 항목: {idx, kind, day, label, entry|None}.
    kind ∈ {header, body, cancelled, holiday, ebs_banner, ebs}.
    행 제외 기능을 위해 idx 부여.
    """
    if isinstance(days, str):
        days = [days]
    out: list[dict] = []
    idx = 0
    for d in days:
        result = build_for_day(df, d)
        out.append(
            {"idx": idx, "kind": "header", "day": d, "entry": result["header_row"],
             "label": f"[{d}요일 헤더] {result['header_row']['values'][COL_A]}"}
        )
        idx += 1
        for entry in result["body"]:
            title = entry["values"][COL_SEASON_TITLE] or ""
            ep = entry["values"][COL_EPISODE_NUMBER] or ""
            out.append({"idx": idx, "kind": "body", "day": d, "entry": entry,
                        "label": f"[{d}] {title} {ep}".strip()})
            idx += 1
        for entry in result["cancelled"]:
            title = entry["values"][COL_SEASON_TITLE] or ""
            out.append({"idx": idx, "kind": "cancelled", "day": d, "entry": entry,
                        "label": f"[{d} 결방] {title}"})
            idx += 1
        for entry in result["holiday_block"]:
            if entry["category"] == "holiday_header":
                out.append({"idx": idx, "kind": "holiday", "day": d, "entry": entry,
                            "label": f"[{d} 연휴지연 헤더] {entry['values'][COL_A]}"})
            else:
                title = entry["values"][COL_SEASON_TITLE] or ""
                out.append({"idx": idx, "kind": "holiday", "day": d, "entry": entry,
                            "label": f"[{d} 연휴지연] {title}"})
            idx += 1
        if d == "금" and result["ebs_section"]:
            out.append({"idx": idx, "kind": "ebs_banner", "day": d, "entry": None,
                        "label": "[EBS 섹션 헤더]"})
            idx += 1
            for entry in result["ebs_section"]:
                title = entry["values"][COL_SEASON_TITLE] or ""
                out.append({"idx": idx, "kind": "ebs", "day": d, "entry": entry,
                            "label": f"[EBS] {title}"})
                idx += 1
    return out


def build_xlsx(df: pd.DataFrame, days: list[str] | str, excluded_indices: set[int] | None = None, gsheet_mode: bool = False) -> bytes:
    """선택한 요일(여러 개 가능)의 검수시트를 한 시트에 이어쓴 .xlsx 바이트.

    - days가 단일 요일 문자열이면 그 요일만, 리스트면 리스트 순서대로 이어 출력.
    - 각 요일마다 [요일 헤더 행 + 본문 + 결방] 순서. 금요일이면 그 뒤에 EBS 배너+섹션.
    """
    if isinstance(days, str):
        days = [days]
    if not days:
        days = ["월"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{days[0]}~{days[-1]}요일" if len(days) > 1 else f"{days[0]}요일"

    # 색상 매핑 선택: 구글시트 호환 모드 vs 기본 (검수시트 템플릿)
    fill_for = _gs_fill_for if gsheet_mode else _fill_for
    over_fill = _GS_FILL_OVER_FINAL if gsheet_mode else _FILL_OVER_FINAL
    banner_fill = _GS_FILL_EBS_BANNER if gsheet_mode else _FILL_EBS_BANNER

    # 시트 상단에 종영 의심 알림 (있으면)
    alerts = check_jongyeong_alerts(df)
    for alert in alerts:
        a_vals = [""] * len(CHECKSHEET_HEADERS)
        a_vals[0] = "확인 필요"
        a_vals[2] = alert["message"]
        ws.append(a_vals)
        row_idx = ws.max_row
        for col_idx in range(1, len(CHECKSHEET_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = _FILL_ALERT
            ws.cell(row=row_idx, column=col_idx).font = _FONT_BASE
        ws.cell(row=row_idx, column=1).font = Font(name="Arial", size=10, bold=True, color="FFC62828")
        ws.cell(row=row_idx, column=COL_NEW_SUMMARY + 1).alignment = Alignment(
            wrap_text=True, vertical="center"
        )

    def write_entry(entry: dict) -> None:
        ws.append(entry["values"])
        row_idx = ws.max_row
        # 모든 셀에 Arial 10pt 강제 적용 (검수시트 템플릿 통일)
        for col_idx in range(1, len(entry["values"]) + 1):
            ws.cell(row=row_idx, column=col_idx).font = _FONT_BASE
        if not entry.get("no_fill"):
            # over_final이면 카테고리 음영을 덮고 행 전체를 옅은 파랑으로
            fill = over_fill if entry.get("over_final") else fill_for(entry["category"])
            if fill is not None:
                for col_idx in range(1, len(entry["values"]) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
        # xlsx에서는 종영 셀에 노란 음영을 입히지 않음 (사용자 룰: 미리보기에만 음영, 다운로드 X)
        if entry.get("jongyeong") and entry["category"] not in ("holiday_header", "holiday_body"):
            ws.cell(row=row_idx, column=1).font = _FONT_BOLD
        if entry["category"] in ("header", "header_holiday", "holiday_header"):
            # holiday_header(연휴지연)만 빨간 굵게로 강조. 요일 헤더(header, header_holiday)는 일반 굵기.
            if entry["category"] == "holiday_header":
                ws.cell(row=row_idx, column=1).font = _FONT_BOLD_RED
            else:
                ws.cell(row=row_idx, column=1).font = _FONT_BASE
            ws.cell(row=row_idx, column=COL_NEW_SUMMARY + 1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )

    excluded = set(excluded_indices or [])
    ordered = build_ordered_entries(df, days)
    for item in ordered:
        if item["idx"] in excluded:
            continue
        if item["kind"] == "ebs_banner":
            banner_vals = [""] * len(CHECKSHEET_HEADERS)
            banner_vals[0] = "EBS"
            ws.append(banner_vals)
            banner_row = ws.max_row
            for col_idx in range(1, len(CHECKSHEET_HEADERS) + 1):
                ws.cell(row=banner_row, column=col_idx).fill = banner_fill
                ws.cell(row=banner_row, column=col_idx).font = _FONT_BASE
            ws.cell(row=banner_row, column=1).font = _FONT_BOLD
        elif item["entry"] is not None:
            write_entry(item["entry"])

    width_map = {
        COL_A + 1: 10,
        COL_TIER + 1: 6,
        COL_SEASON_ID + 1: 11,
        COL_SEASON_CODE + 1: 12,
        COL_SEASON_TITLE + 1: 38,
        COL_MAPPING_TYPE + 1: 11,
        COL_EPISODE_NUMBER + 1: 10,
        COL_FORMAL_NUMBER + 1: 10,
        COL_CP + 1: 18,
        COL_FILE_PATH + 1: 50,
        COL_NEW_SUMMARY + 1: 70,
    }
    for col, w in width_map.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def select_day_range(start: str, end: str) -> list[str]:
    """월~금 고정 순서 안에서 start~end 구간을 추출. start가 end보다 뒤면 swap.
    토/일은 검수시트 대상에서 제외된다 (금요일에 합쳐짐).
    """
    if start not in WEEKDAYS or end not in WEEKDAYS:
        return []
    si, ei = WEEKDAYS.index(start), WEEKDAYS.index(end)
    if si > ei:
        si, ei = ei, si
    return WEEKDAYS[si : ei + 1]
